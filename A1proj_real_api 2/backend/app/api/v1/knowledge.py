from pathlib import Path
import json
from datetime import datetime
from urllib.parse import quote

from fastapi import (
    APIRouter,
    Body,
    Depends,
    File,
    Form,
    Header,
    HTTPException,
    UploadFile,
)
from fastapi.responses import FileResponse

from app.core.config import get_settings
from app.db import init_db, get_session
from app.knowledge.dynamic_store import DynamicKnowledgeStore
from app.knowledge.graph_service import KnowledgeGraphService as JsonGraphService
from app.knowledge.sync_service import KnowledgeSyncService
from app.langchain.vector_store import DashVectorStore
from app.services.graph_db_service import GraphDBService
from app.pipeline import TripleExtractor

router = APIRouter(prefix="/knowledge", tags=["动态知识库"])


def verify_admin(x_admin_token: str | None = Header(default=None)):
    settings = get_settings()
    if not x_admin_token or x_admin_token != settings.ADMIN_TOKEN:
        raise HTTPException(
            status_code=403, detail="需要有效管理员 Token（Header: X-Admin-Token）"
        )


def _rebuild_index():
    """重建 DashVector 向量索引（内部函数）。"""
    svc = KnowledgeSyncService()
    docs = svc.load_all_documents()
    store = DashVectorStore()
    store.save(docs)
    from app.api.v1.chat import reload

    reload()


def _manual_search_dirs() -> list[Path]:
    settings = get_settings()
    svc = KnowledgeSyncService()
    return list(svc._manual_dirs()) + [Path(settings.DATA_DIR) / "_pending"]


def _resolve_manual_path(filename: str) -> tuple[str, Path | None]:
    safe_name = Path(filename).name
    if safe_name != filename:
        raise HTTPException(status_code=400, detail="文件名不合法")
    for d in _manual_search_dirs():
        fp = d / safe_name
        if fp.exists():
            return safe_name, fp
    return safe_name, None


def _manual_doc_counts() -> dict[str, int]:
    svc = KnowledgeSyncService()
    try:
        docs = svc.load_all_documents()
    except Exception:
        docs = []
    counts: dict[str, int] = {}
    for doc in docs:
        source = doc.get("source", "")
        if source:
            counts[source] = counts.get(source, 0) + 1
    return counts


def _diagnose_pdf(filename: str) -> dict:
    """Layered PDF diagnostics used by preview and content extraction."""
    safe_name, file_path = _resolve_manual_path(filename)
    svc = KnowledgeSyncService()
    state = svc._load_sync_state()
    file_hashes = state.get("files", {})
    doc_count = _manual_doc_counts().get(safe_name, 0)
    record = {
        "filename": safe_name,
        "record_exists": safe_name in file_hashes or doc_count > 0,
        "record_md5": file_hashes.get(safe_name),
        "doc_count": doc_count,
        "storage_path": None,
    }
    result = {
        "filename": safe_name,
        "file_exists": False,
        "file_size": 0,
        "file_hash": "",
        "is_pdf_header": False,
        "is_encrypted": False,
        "is_scanned": False,
        "page_count": 0,
        "preview_status": "missing",
        "parse_status": "not_started",
        "error_type": "file_missing",
        "message": "原 PDF 文件不存在",
        "record": record,
        "last_checked_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if file_path is None:
        if record["record_exists"]:
            result["message"] = "文件记录存在，但存储文件缺失"
        return result
    record["storage_path"] = str(file_path)
    result["file_exists"] = True

    if not file_path.is_file():
        result.update({"preview_status": "invalid_path", "error_type": "path_error", "message": "文件路径配置错误"})
        return result
    try:
        size = file_path.stat().st_size
    except OSError:
        result.update({"preview_status": "permission_denied", "error_type": "permission_denied", "message": "文件没有读取权限"})
        return result
    result["file_size"] = size
    if size <= 0:
        result.update({"preview_status": "empty_file", "error_type": "empty_file", "message": "文件大小为 0"})
        return result
    try:
        with file_path.open("rb") as fh:
            header = fh.read(8)
    except OSError:
        result.update({"preview_status": "permission_denied", "error_type": "permission_denied", "message": "文件没有读取权限"})
        return result
    if not header.startswith(b"%PDF-"):
        result.update({"preview_status": "format_mismatch", "error_type": "format_mismatch", "message": "文件格式与扩展名不一致"})
        return result

    result["is_pdf_header"] = True
    try:
        result["file_hash"] = file_md5(file_path)
    except Exception:
        result["file_hash"] = ""

    try:
        from pypdf import PdfReader

        reader = PdfReader(str(file_path))
        result["is_encrypted"] = bool(reader.is_encrypted)
        if reader.is_encrypted:
            result.update({"preview_status": "encrypted", "parse_status": "encrypted", "error_type": "encrypted", "message": "PDF 已加密，需要密码"})
            return result
        page_count = len(reader.pages)
        result["page_count"] = page_count
        sample_text = []
        for page in reader.pages[: min(page_count, 5)]:
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
            if text.strip():
                sample_text.append(text.strip())
        result["preview_status"] = "available"
        if "".join(sample_text).strip():
            result.update({"parse_status": "parsed", "error_type": "", "message": "PDF 可预览，已检测到文字层"})
        else:
            result.update({"parse_status": "scanned", "is_scanned": True, "error_type": "scanned_pdf", "message": "PDF 可预览，但未检测到文字层，可能为扫描件"})
    except Exception as exc:
        result.update({"preview_status": "damaged", "parse_status": "parse_error", "error_type": "damaged_or_parse_error", "message": f"PDF 内容损坏或解析服务异常: {str(exc)[:120]}"})
    return result


# ==================== MySQL 知识图谱 ====================


@router.get("/graph")
def get_full_graph(entity_type: str | None = None):
    """全量图谱（ECharts 格式）。entity_type 可选筛选: device/component/fault/fault_cause/solution。"""
    try:
        svc = GraphDBService()
        return svc.get_full_graph(entity_type)
    except Exception:
        # MySQL 未连接时回退到 JSON 图谱，转换为 ECharts 兼容格式
        graph = JsonGraphService()._load()
        raw_nodes = graph.get("nodes", [])
        # 旧格式 {id, label, type} → 新格式 {id, name, category}
        # 类型名映射：对齐前端 typeFilter 的 entity_type 值
        TYPE_MAP = {
            "device_model": "device",
            "component": "component",
            "fault": "fault",
            "document": "document",
            "source_file": "source_file",
            "tag": "tag",
        }
        nodes = [
            {"id": n["id"], "name": n.get("label", n.get("id", "")),
             "category": TYPE_MAP.get(n.get("type", "document"), n.get("type", "document"))}
            for n in raw_nodes
        ]
        raw_edges = graph.get("edges", [])
        if entity_type:
            keep_ids = {n["id"] for n in nodes if n["category"] == entity_type}
            nodes = [n for n in nodes if n["category"] == entity_type]
            raw_edges = [e for e in raw_edges
                         if e.get("source") in keep_ids or e.get("target") in keep_ids]
        edges = [
            {"source": e.get("source", ""), "relation": e.get("relation", "关联"),
             "target": e.get("target", "")}
            for e in raw_edges
        ]
        cats = list({n["category"] for n in nodes})
        categories = [
            {"name": c, "itemStyle": {"color": "#4a90d9"}} for c in cats
        ]
        return {
            "code": 200,
            "data": {"nodes": nodes, "edges": edges, "categories": categories},
        }


@router.get("/graph/node/{biz_id}")
def get_node_detail(biz_id: str):
    """单个节点详情 + 邻接关系。"""
    svc = GraphDBService()
    detail = svc.get_node_detail(biz_id)
    if detail is None:
        raise HTTPException(status_code=404, detail=f"节点 {biz_id} 不存在")
    return {"code": 200, "data": detail}


@router.get("/graph/stats")
def get_graph_stats():
    """各类型实体与关系数量统计。"""
    try:
        svc = GraphDBService()
        return {"code": 200, "data": svc.get_stats()}
    except Exception:
        graph = JsonGraphService()._load()
        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])
        type_counts = {}
        for n in nodes:
            t = n.get("type", "unknown")
            type_counts[t] = type_counts.get(t, 0) + 1
        type_counts["relation"] = len(edges)
        return {"code": 200, "data": type_counts, "source": "json_fallback"}


@router.post("/graph/extract", dependencies=[Depends(verify_admin)])
def trigger_extraction():
    """触发全量 LLM 三元组抽取（管理员）。"""
    svc = KnowledgeSyncService()
    docs = svc.load_all_documents()
    extractor = TripleExtractor(batch_size=5)
    result = extractor.extract_from_documents(docs)
    return {
        "code": 200,
        "data": {
            "total_docs": result.total_docs,
            "raw_triples": result.raw_triples,
            "valid_triples": result.valid_triples,
            "unique_triples": result.unique_triples,
            "entities_inserted": result.entities_inserted,
            "relations_inserted": result.relations_inserted,
            "errors": result.errors[:10],
        }
    }


@router.post("/graph/extract/{doc_id}", dependencies=[Depends(verify_admin)])
def trigger_single_extraction(doc_id: str):
    """对指定文档触发 LLM 三元组抽取（管理员）。"""
    svc = KnowledgeSyncService()
    docs = svc.load_all_documents()
    target = [d for d in docs if d.get("id") == doc_id]
    if not target:
        raise HTTPException(status_code=404, detail=f"文档 {doc_id} 不存在")
    extractor = TripleExtractor(batch_size=1)
    result = extractor.extract_from_documents(target)
    return {
        "code": 200,
        "data": {
            "raw_triples": result.raw_triples,
            "unique_triples": result.unique_triples,
            "entities_inserted": result.entities_inserted,
            "relations_inserted": result.relations_inserted,
        }
    }


# ==================== Neo4j AuraDB 知识图谱（直连云实例） ====================

from app.services.neo4j_service import Neo4jService as _Neo4jService


@router.get("/neo4j-graph/node/{eid}")
def get_neo4j_node_detail(eid: str):
    """Neo4j AuraDB 单个节点详情。"""
    svc = _Neo4jService()
    try:
        detail = svc.get_node_detail(eid)
        if detail is None:
            raise HTTPException(status_code=404, detail=f"节点 {eid} 不存在")
        return {"code": 200, "data": detail}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Neo4j AuraDB 查询失败: {exc}")


@router.get("/neo4j-graph")
def get_neo4j_graph(entity_type: str | None = None):
    """Neo4j AuraDB 全量图谱（ECharts 格式）。entity_type 可选: device/component/fault/fault_cause/solution。"""
    svc = _Neo4jService()
    try:
        return svc.get_full_graph(entity_type)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Neo4j AuraDB 查询失败: {exc}")


@router.get("/json-graph")
def get_json_graph():
    """JSON 文件图谱（兼容旧接口）。"""
    graph = JsonGraphService()._load()
    return {"nodes": graph.get("nodes", []), "edges": graph.get("edges", [])}


@router.get("/devices")
def get_all_devices():
    graph = JsonGraphService()._load()
    devices = []
    for node in graph.get("nodes", []):
        if node.get("type") == "device_model":
            devices.append({"id": node.get("id"), "name": node.get("label")})
    unique_devices = []
    seen_names = set()
    for device in devices:
        if device["name"] not in seen_names:
            seen_names.add(device["name"])
            unique_devices.append(device)
    return sorted(unique_devices, key=lambda x: x["name"])


# ==================== 手册文件管理 ====================


@router.get(
    "/manuals", summary="获取所有手册文件列表", dependencies=[Depends(verify_admin)]
)
def list_manual_files():
    settings = get_settings()
    sync_service = KnowledgeSyncService()
    state = sync_service._load_sync_state()
    file_hashes = state.get("files", {})
    docs = sync_service.load_all_documents()

    doc_count_by_source = {}
    for doc in docs:
        source = doc.get("source", "")
        doc_count_by_source[source] = doc_count_by_source.get(source, 0) + 1

    files = []
    manual_files = sync_service._list_manual_files()
    cat_overrides = _load_categories()
    for p in manual_files:
        size_kb = round(p.stat().st_size / 1024)
        file_type = p.suffix.lower().replace(".", "").upper()
        pdf_status = {}  # skip per-file diagnose for list performance
        category = cat_overrides.get(p.name) or (
            "总装设备检修手册" if "报警代码" in p.name
            else ("机床设备维修手册" if p.name in file_hashes else "未分类")
        )
        files.append(
            {
                "filename": p.name,
                "size_kb": size_kb,
                "type": file_type,
                "md5": file_hashes.get(p.name, "未同步"),
                "doc_count": doc_count_by_source.get(p.name, 0),
                "status": "已同步" if p.name in file_hashes else "待同步",
                "category": category,
                "preview_status": pdf_status.get("preview_status", "available"),
                "parse_status": pdf_status.get("parse_status", "parsed" if doc_count_by_source.get(p.name, 0) else "not_started"),
                "is_scanned": pdf_status.get("is_scanned", False),
                "is_encrypted": pdf_status.get("is_encrypted", False),
                "diagnostic_message": pdf_status.get("message", ""),
            }
        )
    return sorted(files, key=lambda x: x["filename"])

# 分类持久化
_categories_path = Path(get_settings().KNOWLEDGE_DIR) / "manual_categories.json"

def _load_categories() -> dict:
    if _categories_path.exists():
        try:
            return json.loads(_categories_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _save_categories(data: dict):
    _categories_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

@router.put("/manuals/{filename}/category", dependencies=[Depends(verify_admin)])
def update_manual_category(filename: str, body: dict = Body(...)):
    cats = _load_categories()
    cats[filename] = body.get("category", "")
    _save_categories(cats)
    return {"success": True}

@router.post(
    "/manuals/upload", summary="上传手册文件", dependencies=[Depends(verify_admin)]
)
async def upload_manual(file: UploadFile = File(...), auto_sync: bool = False):
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")
    settings = get_settings()
    data_dir = Path(settings.DATA_DIR)
    data_dir.mkdir(parents=True, exist_ok=True)
    save_path = data_dir / file.filename
    content = await file.read()
    save_path.write_bytes(content)
    result = None
    if auto_sync:
        svc = KnowledgeSyncService()
        result = svc.sync()
        _rebuild_index()
    return {
        "success": True,
        "filename": file.filename,
        "size_kb": round(len(content) / 1024, 2),
        "auto_sync": auto_sync,
        "sync_result": result,
    }


@router.get("/manuals/{filename:path}/content", dependencies=[Depends(verify_admin)])
def get_manual_content(filename: str):
    """读取手册文件内容（支持 PDF/DOCX/TXT）。"""
    safe_name, file_path = _resolve_manual_path(filename)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"文件 {safe_name} 不存在或尚未入库")

    suffix = file_path.suffix.lower()
    content = ""
    try:
        if suffix == ".txt":
            content = file_path.read_text(encoding="utf-8")[:50000]
        elif suffix == ".pdf":
            diagnosis = _diagnose_pdf(safe_name)
            if diagnosis["preview_status"] in {"missing", "invalid_path", "permission_denied", "empty_file", "format_mismatch", "encrypted", "damaged"}:
                raise HTTPException(status_code=422, detail=diagnosis["message"])
            content = ""
            for lib in ("pypdf", "fitz"):
                try:
                    if lib == "pypdf":
                        from pypdf import PdfReader
                        reader = PdfReader(str(file_path))
                        pages = [p.extract_text() or "" for p in reader.pages]
                    else:
                        import fitz
                        doc = fitz.open(str(file_path))
                        pages = [p.get_text() for p in doc]
                    content = "\n\n".join(pages)[:50000]
                    if len(content.strip()) > 50:
                        break
                except Exception:
                    continue
            if not content.strip():
                content = "[PDF 可以查看，但未检测到文字层，可能为扫描件；原文件预览不受影响。]"
        elif suffix in (".doc", ".docx"):
            try:
                from docx import Document
                doc = Document(str(file_path))
                content = "\n".join(p.text for p in doc.paragraphs)[:50000]
            except Exception:
                content = "[无法解析 DOCX 文件，请确认格式正确]"
        elif suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(file_path), read_only=True, data_only=True)
                rows = []
                for ws in wb.worksheets[:3]:
                    rows.append(f"【工作表：{ws.title}】")
                    for row in ws.iter_rows(max_row=120, values_only=True):
                        values = ["" if v is None else str(v) for v in row]
                        if any(v.strip() for v in values):
                            rows.append(" | ".join(values))
                    rows.append("")
                content = "\n".join(rows)[:50000] or "[Excel 文件为空]"
            except Exception as exc:
                raise HTTPException(status_code=422, detail=f"Excel 解析失败: {exc}")
        elif suffix == ".xls":
            try:
                import xlrd
                book = xlrd.open_workbook(str(file_path))
                rows = []
                for sheet in book.sheets()[:3]:
                    rows.append(f"【工作表：{sheet.name}】")
                    for r in range(min(sheet.nrows, 120)):
                        values = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
                        if any(values):
                            rows.append(" | ".join(values))
                    rows.append("")
                content = "\n".join(rows)[:50000] or "[Excel 文件为空]"
            except Exception as exc:
                raise HTTPException(status_code=422, detail=f"Excel 解析失败: {exc}")
        else:
            content = f"[不支持预览 .{suffix} 格式]"
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(status_code=422, detail=f"读取文件失败: {e}")

    return {"filename": safe_name, "content": content, "size_kb": round(file_path.stat().st_size / 1024)}


@router.get("/manuals/{filename:path}/diagnostics", dependencies=[Depends(verify_admin)])
def get_manual_diagnostics(filename: str):
    safe_name = Path(filename).name
    if Path(safe_name).suffix.lower() != ".pdf":
        _, file_path = _resolve_manual_path(filename)
        if file_path is None:
            raise HTTPException(status_code=404, detail=f"文件 {safe_name} 不存在或尚未入库")
        return {
            "filename": safe_name,
            "file_exists": True,
            "file_size": file_path.stat().st_size,
            "preview_status": "available",
            "parse_status": "not_pdf",
            "message": "非 PDF 文件，使用文本预览",
            "last_checked_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    return _diagnose_pdf(filename)


@router.get("/manuals/{filename:path}/raw", dependencies=[Depends(verify_admin)])
def get_manual_raw(filename: str):
    safe_name, file_path = _resolve_manual_path(filename)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"文件 {safe_name} 不存在或尚未入库")
    if file_path.suffix.lower() != ".pdf":
        raise HTTPException(status_code=415, detail="当前原文件预览仅支持 PDF")
    diagnosis = _diagnose_pdf(safe_name)
    if diagnosis["preview_status"] in {"missing", "invalid_path", "permission_denied", "empty_file", "format_mismatch"}:
        raise HTTPException(status_code=422, detail=diagnosis["message"])
    return FileResponse(
        path=str(file_path),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{quote(safe_name)}"},
    )


@router.delete("/manuals/{filename}", dependencies=[Depends(verify_admin)])
def delete_manual(filename: str):
    """删除手册文件及对应知识库文档。"""
    svc = KnowledgeSyncService()
    safe_name = Path(filename).name
    file_path = next((d / safe_name for d in svc._manual_dirs() if (d / safe_name).exists()), None)
    if file_path is None or not file_path.exists():
        raise HTTPException(status_code=404, detail=f"文件 {safe_name} 不存在")
    file_path.unlink()
    # 从知识库中移除该文件的所有文档
    docs = svc.load_all_documents()
    before = len(docs)
    docs = [d for d in docs if d.get("source") != safe_name]
    after = len(docs)
    svc.knowledge_path.write_text(json.dumps(docs, ensure_ascii=False, indent=2), encoding="utf-8")
    # 更新同步状态
    state = svc._load_sync_state()
    if safe_name in state.get("files", {}):
        del state["files"][safe_name]
        svc._save_sync_state(state)
    return {"success": True, "filename": safe_name, "removed_docs": before - after}


@router.post("/sync", dependencies=[Depends(verify_admin)])
def trigger_sync():
    """触发全量知识库同步。"""
    svc = KnowledgeSyncService()
    result = svc.sync()
    _rebuild_index()
    return {"success": True, "result": result}


@router.post("/sync/{filename}", dependencies=[Depends(verify_admin)])
def trigger_single_sync(filename: str):
    """单独同步一个手册文件。"""
    safe_name = Path(filename).name
    svc = KnowledgeSyncService()
    file_path = next((d / safe_name for d in svc._manual_dirs() if (d / safe_name).exists()), None)
    if file_path is None or not file_path.exists():
        raise HTTPException(status_code=404, detail=f"文件 {safe_name} 不存在")
    try:
        from app.knowledge.document_parser import parse_manual_file, assign_ids
        raw_docs = parse_manual_file(file_path)
        docs = assign_ids(raw_docs, safe_name)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)[:200]}")
    # 移除旧条目 + 合并新条目
    all_docs = svc.load_all_documents()
    all_docs = [d for d in all_docs if d.get("source") != safe_name]
    all_docs.extend(docs)
    svc.knowledge_path.write_text(json.dumps(all_docs, ensure_ascii=False, indent=2), encoding="utf-8")
    # 更新同步状态
    state = svc._load_sync_state()
    state["files"][safe_name] = file_md5(file_path)
    state["document_count"] = len(all_docs)
    state["manual_count"] = len([d for d in all_docs if d.get("source")])
    import time
    state["updated_at"] = time.time()
    svc._save_sync_state(state)
    _rebuild_index()
    return {"success": True, "filename": safe_name, "added_docs": len(docs),
            "total_docs": len(all_docs)}


from app.knowledge.document_parser import file_md5

# ==================== 手册上传申请流程 ===


@router.post("/manuals/request", summary="提交手册录入申请")
async def request_manual_upload(
    file: UploadFile = File(...),
    username: str = Form(...),
    description: str = Form(""),
    device_model: str | None = Form(None),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".pdf", ".docx"}:
        raise HTTPException(status_code=400, detail="仅支持 PDF 和 DOCX 格式")
    settings = get_settings()
    temp_dir = Path(settings.DATA_DIR) / "_pending"
    temp_dir.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    save_path = temp_dir / file.filename
    save_path.write_bytes(content)
    store = DynamicKnowledgeStore()
    req = store.add_manual_request(
        filename=file.filename,
        file_size=len(content),
        applicant=username,
        description=description,
        device_model=device_model,
    )
    return {"success": True, "request": req}


@router.get(
    "/manuals/requests",
    summary="获取手册申请列表",
    dependencies=[Depends(verify_admin)],
)
def list_manual_requests(status: str | None = None):
    store = DynamicKnowledgeStore()
    return {"requests": store.list_manual_requests(status=status)}


@router.post(
    "/manuals/requests/{request_id}/review",
    summary="审核手册申请",
    dependencies=[Depends(verify_admin)],
)
def review_manual_request(
    request_id: str,
    approve: bool = Body(..., embed=True),
    reviewer: str = Body("admin", embed=True),
    comment: str = Body("", embed=True),
    auto_sync: bool = Body(True, embed=True),
):
    store = DynamicKnowledgeStore()
    settings = get_settings()
    temp_dir = Path(settings.DATA_DIR) / "_pending"
    target_dir = Path(settings.DATA_DIR)
    req = store.review_manual_request(
        request_id, approve=approve, reviewer=reviewer, comment=comment
    )
    if approve:
        src = temp_dir / req["filename"]
        dst = target_dir / req["filename"]
        if src.exists():
            import shutil

            shutil.move(str(src), str(dst))
            if auto_sync:
                svc = KnowledgeSyncService()
                result = svc.sync()
                _rebuild_index()
                return {"success": True, "request": req, "sync_result": result}
    return {"success": True, "request": req}
